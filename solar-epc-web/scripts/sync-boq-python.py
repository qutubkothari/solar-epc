"""
BOQ workbook sync — Python/SQLite fallback (bypasses Prisma engine).
Mirrors the logic in sync-boq-workbooks.js exactly.
"""
import sqlite3
import re
import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl not found, trying to install...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.normpath(os.path.join(SCRIPT_DIR, "..", ".."))
BOQ_LIST_PATH     = os.path.join(REPO_ROOT, "BOQ LIST.xlsx")
BOQ_SEQUENCE_PATH = os.path.join(REPO_ROOT, "BOQ Sequence.xlsx")
DB_PATH           = os.path.join(SCRIPT_DIR, "..", "prisma", "dev.db")

# ── Helpers ──────────────────────────────────────────────────────────────────
def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()

def normalize(value):
    return clean(value).lower()

def to_number(value):
    if value is None or value == "":
        return 0.0
    try:
        s = re.sub(r"[^0-9.+\-]", "", str(value))
        return float(s) if s else 0.0
    except ValueError:
        return 0.0

def pricing_unit_from_uom(uom):
    unit = clean(uom).upper()
    if unit == "WP":
        return "RS_PER_WATT"
    if unit == "KW":
        return "RS_PER_KW"
    return "PER_UNIT"

def deterministic_sku(item_head, item_type, rating, make):
    parts = [item_head, item_type, rating, make]
    slugs = [re.sub(r"[^a-z0-9]+", "-", normalize(p)).strip("-") for p in parts]
    return "__".join(s for s in slugs if s)

# ── Read Sequence workbook ────────────────────────────────────────────────────
def get_sequence_heads():
    wb = openpyxl.load_workbook(BOQ_SEQUENCE_PATH, data_only=True)
    # Try Sheet2 first (per prior inspection), fall back to first sheet
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.worksheets[0]
    heads = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        head = clean(row[1]) if len(row) > 1 else ""
        if head:
            heads.append(head)
    return heads

# ── Read BOQ LIST workbook ────────────────────────────────────────────────────
def get_boq_items(valid_heads):
    wb = openpyxl.load_workbook(BOQ_LIST_PATH, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if len(row) < 7:
            continue

        item_head   = clean(row[1])
        raw_type    = clean(row[2])
        raw_rating  = clean(row[3])
        make        = clean(row[4])
        details     = clean(row[5])
        uom         = clean(row[6])
        rate        = to_number(row[7]) if len(row) > 7 else 0.0
        gst_raw     = to_number(row[8]) if len(row) > 8 else 0.0
        warranty    = clean(row[9]) if len(row) > 9 else ""

        item_type = raw_type or raw_rating or item_head
        rating    = raw_rating if raw_type else ""

        if item_head not in valid_heads or not item_type:
            continue

        name = f"{item_type} - {rating}" if rating else item_type

        desc_parts = []
        if rating:
            desc_parts.append(f"Rating/Capacity: {rating}")
        if details:
            desc_parts.append(details)
        if warranty:
            desc_parts.append(f"Warranty: {warranty}")
        description = " | ".join(desc_parts) or None

        tax_percent = gst_raw * 100 if gst_raw <= 1 else gst_raw

        items.append({
            "itemHead":    item_head,
            "itemType":    item_type,
            "rating":      rating,
            "name":        name,
            "brand":       make or None,
            "description": description,
            "uom":         uom or "NOS",
            "unitPrice":   rate,
            "taxPercent":  tax_percent,
            "pricingUnit": pricing_unit_from_uom(uom),
            "sku":         deterministic_sku(item_head, item_type, rating, make),
            "category":    item_head,
        })
    return items

# ── Sync to SQLite ────────────────────────────────────────────────────────────
def sync():
    print(f"Reading sequence heads from: {BOQ_SEQUENCE_PATH}")
    valid_heads = get_sequence_heads()
    print(f"  Found {len(valid_heads)} valid heads: {valid_heads[:5]}...")

    print(f"Reading BOQ items from: {BOQ_LIST_PATH}")
    boq_items = get_boq_items(valid_heads)
    print(f"  Found {len(boq_items)} BOQ items to sync")

    if not boq_items:
        print("ERROR: No BOQ items found — aborting sync")
        sys.exit(1)

    db_path = os.path.normpath(DB_PATH)
    print(f"Connecting to SQLite: {db_path}")
    con = sqlite3.connect(db_path)
    cur = con.cursor()

    # Deactivate all items
    cur.execute("UPDATE Item SET isActive = 0")
    deactivated = cur.rowcount
    print(f"  Deactivated {deactivated} existing items")

    created = 0
    updated = 0

    for item in boq_items:
        cur.execute("SELECT id FROM Item WHERE sku = ?", (item["sku"],))
        row = cur.fetchone()
        if row:
            existing_id = row[0]
            cur.execute("""
                UPDATE Item SET
                    name = ?, description = ?, brand = ?,
                    unitPrice = ?, taxPercent = ?, marginPercent = 0,
                    uom = ?, pricingUnit = ?, category = ?, isActive = 1
                WHERE id = ?
            """, (
                item["name"], item["description"], item["brand"],
                item["unitPrice"], item["taxPercent"],
                item["uom"], item["pricingUnit"], item["category"],
                existing_id,
            ))
            updated += 1
        else:
            import uuid
            new_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO Item
                    (id, name, description, brand, unitPrice, taxPercent,
                     marginPercent, uom, pricingUnit, category, sku, isActive,
                     createdAt, updatedAt)
                VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, 1,
                        datetime('now'), datetime('now'))
            """, (
                new_id,
                item["name"], item["description"], item["brand"],
                item["unitPrice"], item["taxPercent"],
                item["uom"], item["pricingUnit"], item["category"],
                item["sku"],
            ))
            created += 1

    con.commit()
    con.close()
    print(f"\nBOQ sync complete. Created: {created}, Updated: {updated}, Active: {len(boq_items)}")

if __name__ == "__main__":
    sync()
